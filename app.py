import streamlit as st
import gspread
import pandas as pd
import datetime
import random
import time
import json
from collections import defaultdict
import io
import plotly.express as px
import plotly.graph_objects as go
import re

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

st.set_page_config(
    page_title="🔍 RADAR DE DIAGNÓSTICO EN VIVO",
    layout="wide"
)

st.title("🔍 RADAR DE ERROR: No vamos a adivinar más")

st.warning("""
Este es un diagnóstico en vivo. 
**Haz lo siguiente:**
1. Inicia sesión como usuario normal.
2. Cuando se cargue la página, mira el mensaje que aparecerá justo debajo de esta advertencia.
3. **Copia y pega ese mensaje aquí.**
""")

# ========== LOGIN ==========
if 'df_personal' not in st.session_state:
    # Simulamos la carga básica para que puedas loguearte
    try:
        creds = st.secrets["gsheets"]
        gc = gspread.service_account_from_dict(creds)
        sh = gc.open_by_key("1TH32e7TkB4RYEKxxRhGRnnYxjkT9AF9TeuplxtYih1Y")
        ws_personal = sh.worksheet("Personal")
        all_values = ws_personal.get_all_values()
        header = [str(col).strip().upper() for col in all_values[0]]
        data = all_values[1:]
        st.session_state.df_personal = pd.DataFrame(data, columns=header)
        ws_usuarios = sh.worksheet("Usuarios")
        all_users = ws_usuarios.get_all_values()
        header_users = [str(col).strip().upper() for col in all_users[0]]
        data_users = all_users[1:]
        st.session_state.df_usuarios = pd.DataFrame(data_users, columns=header_users)
    except:
        pass

if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False

if st.session_state.logged_in:
    user = st.session_state.user_data
    
    with st.sidebar:
        st.markdown("### 👤 Panel de Usuario")
        st.markdown(f"**Nombre:** {user['NOMBRE']}")
        st.markdown(f"**📄 DNI:** {user['DNI']}")
        st.markdown(f"**🏢 Dependencia:** {user['DEPENDENCIA']}")
        if st.button("🚪 Cerrar Sesión"):
            st.session_state.logged_in = False
            st.rerun()
    
    # === EL RADAR DE DIAGNÓSTICO ===
    st.subheader("🧐 ¿Qué está pasando con los datos cuando entras como usuario?")
    
    # Verificamos 3 cosas clave
    datos_filtrados = st.session_state.df_personal[st.session_state.df_personal['DEPENDENCIA'].astype(str).str.lower() == user['DEPENDENCIA'].lower()].copy()
    
    st.write("**1. ¿Cuántos registros tiene tu dependencia?**")
    st.code(f"Total de registros encontrados: {len(datos_filtrados)}")
    
    st.write("**2. ¿Están las columnas esenciales?**")
    columnas = datos_filtrados.columns.tolist()
    st.code(f"Columnas disponibles: {columnas}")
    
    st.write("**3. ¿Qué datos se están mostrando (Primera fila)?**")
    if len(datos_filtrados) > 0:
        st.json(datos_filtrados.iloc[0].to_dict())
    else:
        st.error("❌ NO HAY DATOS PARA MOSTRAR. No se encontró ningún registro.")
    
    st.success("✅ Diagnóstico completado. Ahora dime qué ves en los 3 puntos de arriba.")

else:
    col1, col2 = st.columns([1, 1])
    with col1:
        st.markdown("### Iniciar Sesión")
        dni_input = st.text_input("📄 DNI", placeholder="Ingrese su DNI")
        clave_input = st.text_input("🔒 Clave", type="password", placeholder="Ingrese su contraseña")
        if st.button("🚪 Ingresar", type="primary", use_container_width=True):
            if dni_input and clave_input:
                df_usuarios = st.session_state.df_usuarios
                usuario = df_usuarios[
                    (df_usuarios['DNI'].astype(str).str.lower() == dni_input.lower()) &
                    (df_usuarios['CLAVE'].astype(str).str.lower() == clave_input.lower())
                ]
                if not usuario.empty:
                    st.session_state.logged_in = True
                    st.session_state.user_data = usuario.iloc[0]
                    st.rerun()
                else:
                    st.error("❌ DNI o Clave incorrectos")
            else:
                st.warning("⚠️ Complete ambos campos")
